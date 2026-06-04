from datetime import date, datetime
from decimal import Decimal
from flask import (render_template, redirect, url_for, flash,
                   request, current_app, send_file)
from flask_login import login_required, current_user
from sqlalchemy import or_, extract
import io

from app.extensions import db
from app.models import (CuentaCaja, MovimientoCaja, PagoAlumno,
                        LiquidacionProfesor, LiquidacionItem, AjusteCaja,
                        Inscripcion, Alumno, Profesor, Usuario, Clase, Asistencia,
                        CUENTA_EFECTIVO, CUENTA_TRANSFERENCIA, CUENTA_OTRO,
                        CAT_CUOTA, CAT_COBRO_EXTRA, CAT_ING_EXTRA,
                        CAT_LIQ_PROFESOR, CAT_GASTO_OTRO,
                        ASIST_PRESENTE)
from app.caja.forms import (PagoAlumnoForm, MovimientoExtraForm,
                             AjusteCajaForm, LiquidacionForm, PagarLiquidacionForm)
from app.auth.routes import rol_requerido
from . import caja_bp

PORCENTAJE_PROFESOR = 0.70


def _cuenta_por_tipo(tipo_pago):
    """Retorna la CuentaCaja según el tipo de pago."""
    return CuentaCaja.query.filter_by(nombre=tipo_pago).first()


def _registrar_movimiento(cuenta_id, tipo, categoria, descripcion,
                           monto, fecha, referencia_id=None,
                           referencia_tipo=None):
    """Helper para crear un MovimientoCaja."""
    m = MovimientoCaja(
        cuenta_id       = cuenta_id,
        tipo            = tipo,
        categoria       = categoria,
        descripcion     = descripcion,
        monto           = monto,
        fecha           = fecha,
        referencia_id   = referencia_id,
        referencia_tipo = referencia_tipo,
        creado_por_id   = current_user.id,
    )
    db.session.add(m)
    db.session.flush()
    return m


# ------------------------------------------------------------------
# Dashboard caja
# ------------------------------------------------------------------
@caja_bp.route('/')
@login_required
@rol_requerido('admin')
def index():
    hoy   = date.today()
    mes   = request.args.get('mes',  hoy.month, type=int)
    anio  = request.args.get('anio', hoy.year,  type=int)

    cuentas = CuentaCaja.query.filter_by(activo=True).all()

    # Movimientos del mes
    movimientos_mes = (MovimientoCaja.query
                       .filter(extract('month', MovimientoCaja.fecha) == mes)
                       .filter(extract('year',  MovimientoCaja.fecha) == anio)
                       .order_by(MovimientoCaja.fecha.desc(),
                                 MovimientoCaja.creado_en.desc())
                       .all())

    ingresos_mes = sum(float(m.monto) for m in movimientos_mes if m.tipo == 'ingreso')
    egresos_mes  = sum(float(m.monto) for m in movimientos_mes if m.tipo == 'egreso')

    # Liquidaciones pendientes
    liq_pendientes = (LiquidacionProfesor.query
                      .filter_by(estado='generada')
                      .all())

    return render_template('caja/index.html',
                           cuentas=cuentas,
                           movimientos_mes=movimientos_mes,
                           ingresos_mes=ingresos_mes,
                           egresos_mes=egresos_mes,
                           liq_pendientes=liq_pendientes,
                           mes=mes, anio=anio, hoy=hoy)


# ------------------------------------------------------------------
# Registrar pago de alumno
# ------------------------------------------------------------------
@caja_bp.route('/pago-alumno', methods=['GET', 'POST'])
@login_required
@rol_requerido('admin')
def pago_alumno():
    form = PagoAlumnoForm()

    inscripciones = (Inscripcion.query
                     .join(Alumno).join(Usuario)
                     .filter(Inscripcion.activo == True)
                     .order_by(Usuario.nombre).all())
    form.inscripcion_id.choices = [
        (i.id, f"{i.alumno.usuario.nombre} — {i.curso.nombre}")
        for i in inscripciones
    ]

    # Dict de aranceles para el JS (monto sugerido)
    aranceles = {i.id: round(float(i.arancel_final)) for i in inscripciones}

    if request.method == 'GET':
        form.fecha_pago.data   = date.today()
        form.periodo_mes.data  = date.today().month
        form.periodo_anio.data = date.today().year

    if form.validate_on_submit():
        inscripcion = Inscripcion.query.get(form.inscripcion_id.data)
        cuenta = _cuenta_por_tipo(form.tipo_pago.data)

        descripcion = (f"Cuota {form.periodo_mes.data}/{form.periodo_anio.data} — "
                       f"{inscripcion.alumno.usuario.nombre} / {inscripcion.curso.nombre}")
        mov = _registrar_movimiento(
            cuenta_id       = cuenta.id,
            tipo            = 'ingreso',
            categoria       = CAT_CUOTA,
            descripcion     = descripcion,
            monto           = form.monto.data,
            fecha           = form.fecha_pago.data,
            referencia_tipo = 'pago_alumno',
        )

        pago = PagoAlumno(
            inscripcion_id  = form.inscripcion_id.data,
            periodo_mes     = form.periodo_mes.data,
            periodo_anio    = form.periodo_anio.data,
            monto           = form.monto.data,
            tipo_pago       = form.tipo_pago.data,
            fecha_pago      = form.fecha_pago.data,
            comprobante_nro = form.comprobante_nro.data or None,
            movimiento_id   = mov.id,
        )
        db.session.add(pago)
        db.session.flush()
        mov.referencia_id = pago.id
        db.session.commit()

        _resolver_pendientes_profesor(pago)

        flash(f'Pago registrado correctamente. Comprobante #{pago.id}', 'success')
        return redirect(url_for('caja.comprobante_pago', pago_id=pago.id))

    return render_template('caja/pago_alumno.html', form=form, aranceles=aranceles)




def _resolver_pendientes_profesor(pago):
    """
    Si existe un LiquidacionItem pendiente para esta inscripción,
    lo vincula al pago y actualiza la liquidación.
    """
    items_pendientes = (LiquidacionItem.query
                        .filter_by(inscripcion_id=pago.inscripcion_id,
                                   pago_alumno_id=None)
                        .all())
    for item in items_pendientes:
        item.pago_alumno_id = pago.id
        liq = item.liquidacion

        # Mover monto de pendiente a confirmado
        liq.monto_confirmado = (float(liq.monto_confirmado) +
                                float(item.monto_calculado))
        liq.monto_pendiente  = max(0, float(liq.monto_pendiente) -
                                   float(item.monto_calculado))

        # Si la liquidación ya fue pagada, generar movimiento adicional
        if liq.estado == 'pagada':
            cuenta = _cuenta_por_tipo(liq.tipo_pago or 'efectivo')
            _registrar_movimiento(
                cuenta_id       = cuenta.id,
                tipo            = 'egreso',
                categoria       = CAT_LIQ_PROFESOR,
                descripcion     = (f"Pago pendiente resuelto — "
                                   f"{liq.profesor.usuario.nombre} "
                                   f"{liq.periodo_str}"),
                monto           = item.monto_calculado,
                fecha           = date.today(),
                referencia_id   = liq.id,
                referencia_tipo = 'liquidacion',
            )
    db.session.commit()


# ------------------------------------------------------------------
# Calcular y generar liquidación del profesor
# ------------------------------------------------------------------
@caja_bp.route('/liquidacion/nueva', methods=['GET', 'POST'])
@login_required
@rol_requerido('admin')
def nueva_liquidacion():
    form = LiquidacionForm()
    form.profesor_id.choices = [
        (p.id, p.usuario.nombre)
        for p in Profesor.query.join(Profesor.usuario)
                               .filter_by(activo=True)
                               .order_by(Usuario.nombre).all()
    ]

    if request.method == 'GET':
        form.periodo_mes.data  = date.today().month
        form.periodo_anio.data = date.today().year

    preview = None
    if form.validate_on_submit():
        preview = _calcular_liquidacion(
            form.profesor_id.data,
            form.periodo_mes.data,
            form.periodo_anio.data,
        )

    return render_template('caja/nueva_liquidacion.html',
                           form=form, preview=preview)


# confirmar_liquidacion
def _calcular_liquidacion(profesor_id, mes, anio):
    profesor = Profesor.query.get(profesor_id)
    detalle = []
    total_confirmado = Decimal('0')
    total_pendiente  = Decimal('0')
 
    # ------------------------------------------------------------------
    # PARTE 1 — Clases de sus propios cursos
    # ------------------------------------------------------------------
    for curso in profesor.cursos:
        if not curso.activo:
            continue
 
        clases_mes = (Clase.query
                      .filter_by(curso_id=curso.id)
                      .filter(extract('month', Clase.fecha) == mes)
                      .filter(extract('year',  Clase.fecha) == anio)
                      .filter(Clase.estado != 'cancelada')
                      .all())
        total_clases = len(clases_mes)
        if not total_clases:
            continue
 
        for inscripcion in curso.alumnos_activos():
            # Presencias del profesor ORIGINAL
            # (excluir clases donde fue cubierto por otro)
            presencias = sum(
                1 for c in clases_mes
                for a in c.asistencias
                if a.inscripcion_id == inscripcion.id
                and a.asistencia == ASIST_PRESENTE
                and (c.profesor_reprog_id is None or
                     c.profesor_reprog_id == profesor_id)
            )
 
            # Clases que NO dio porque fueron cubiertas por otro
            clases_no_dictadas = sum(
                1 for c in clases_mes
                if c.profesor_reprog_id is not None
                and c.profesor_reprog_id != profesor_id
            )
 
            if presencias == 0 and clases_no_dictadas == 0:
                continue
 
            valor_por_clase = float(inscripcion.arancel_final) / total_clases
 
            # Monto por clases que sí dictó
            monto = Decimal('0')
            if presencias > 0:
                monto = Decimal(str(
                    round(presencias * valor_por_clase * PORCENTAJE_PROFESOR, 2)
                ))
 
            # Descuento por clases no dictadas (cubiertas por otro)
            descuento = Decimal('0')
            if clases_no_dictadas > 0:
                # Verificar presencia del alumno en esas clases
                presencias_no_dictadas = sum(
                    1 for c in clases_mes
                    for a in c.asistencias
                    if a.inscripcion_id == inscripcion.id
                    and a.asistencia == ASIST_PRESENTE
                    and c.profesor_reprog_id is not None
                    and c.profesor_reprog_id != profesor_id
                )
                if presencias_no_dictadas > 0:
                    descuento = Decimal(str(
                        round(presencias_no_dictadas * valor_por_clase * PORCENTAJE_PROFESOR, 2)
                    ))
 
            monto_neto = monto - descuento
            if monto_neto == 0 and descuento == 0:
                continue
 
            pago = PagoAlumno.query.filter_by(
                inscripcion_id=inscripcion.id,
                periodo_mes=mes,
                periodo_anio=anio,
            ).first()
 
            if pago:
                total_confirmado += monto_neto
            else:
                total_pendiente += monto_neto
 
            detalle.append({
                'tipo':             'propio',
                'curso':            curso.nombre,
                'alumno':           inscripcion.alumno.usuario.nombre,
                'inscripcion_id':   inscripcion.id,
                'clases_dadas':     total_clases,
                'presencias':       presencias,
                'clases_cubiertas': clases_no_dictadas,
                'descuento':        descuento,
                'arancel':          inscripcion.arancel_final,
                'monto':            monto_neto,
                'pago':             pago,
                'pendiente':        pago is None,
            })
 
    # ------------------------------------------------------------------
    # PARTE 2 — Clases extras cubiertas en cursos de otros profesores
    # ------------------------------------------------------------------
    clases_cubiertas = (Clase.query
                        .filter(Clase.profesor_reprog_id == profesor_id)
                        .filter(extract('month', Clase.fecha) == mes)
                        .filter(extract('year',  Clase.fecha) == anio)
                        .filter(Clase.estado != 'cancelada')
                        .all())
 
    for clase in clases_cubiertas:
        curso = clase.curso
 
        # Total clases del mes de ese curso
        total_clases_curso = (Clase.query
                              .filter_by(curso_id=curso.id)
                              .filter(extract('month', Clase.fecha) == mes)
                              .filter(extract('year',  Clase.fecha) == anio)
                              .filter(Clase.estado != 'cancelada')
                              .count())
        if not total_clases_curso:
            continue
 
        for asist in clase.asistencias:
            if asist.asistencia != ASIST_PRESENTE:
                continue
 
            inscripcion = asist.inscripcion
            valor_por_clase = float(inscripcion.arancel_final) / total_clases_curso
            monto = Decimal(str(
                round(valor_por_clase * PORCENTAJE_PROFESOR, 2)
            ))
 
            pago = PagoAlumno.query.filter_by(
                inscripcion_id=inscripcion.id,
                periodo_mes=mes,
                periodo_anio=anio,
            ).first()
 
            if pago:
                total_confirmado += monto
            else:
                total_pendiente += monto
 
            detalle.append({
                'tipo':           'cubierta',
                'curso':          curso.nombre,
                'alumno':         inscripcion.alumno.usuario.nombre,
                'inscripcion_id': inscripcion.id,
                'fecha_clase':    clase.fecha.strftime('%d/%m/%Y'),
                'clases_dadas':   total_clases_curso,
                'presencias':     1,
                'arancel':        inscripcion.arancel_final,
                'monto':          monto,
                'pago':           pago,
                'pendiente':      pago is None,
            })
 
    return {
        'profesor':         profesor,
        'mes':              mes,
        'anio':             anio,
        'detalle':          detalle,
        'total_confirmado': total_confirmado,
        'total_pendiente':  total_pendiente,
    }

@caja_bp.route('/liquidacion/confirmar', methods=['POST'])
@login_required
@rol_requerido('admin')
def confirmar_liquidacion():
    profesor_id  = request.form.get('profesor_id',  type=int)
    periodo_mes  = request.form.get('periodo_mes',  type=int)
    periodo_anio = request.form.get('periodo_anio', type=int)

    existe = LiquidacionProfesor.query.filter_by(
        profesor_id  = profesor_id,
        periodo_mes  = periodo_mes,
        periodo_anio = periodo_anio,
    ).first()
    if existe:
        flash('Ya existe una liquidación para este profesor y período.', 'warning')
        return redirect(url_for('caja.ver_liquidacion', liq_id=existe.id))

    preview = _calcular_liquidacion(profesor_id, periodo_mes, periodo_anio)
    if not preview['detalle']:              # <-- era preview['items']
        flash('No hay clases asistidas para liquidar en este período.', 'warning')
        return redirect(url_for('caja.nueva_liquidacion'))

    liq = LiquidacionProfesor(
        profesor_id      = profesor_id,
        periodo_mes      = periodo_mes,
        periodo_anio     = periodo_anio,
        monto_confirmado = preview['total_confirmado'],
        monto_pendiente  = preview['total_pendiente'],
        estado           = 'generada',
    )
    db.session.add(liq)
    db.session.flush()

    for item in preview['detalle']:         # <-- era preview['items']
        li = LiquidacionItem(
            liquidacion_id   = liq.id,
            inscripcion_id   = item['inscripcion_id'],
            clases_dadas     = item['clases_dadas'],
            clases_asistidas = item['presencias'],
            arancel_acordado = item['arancel'],
            monto_calculado  = item['monto'],
            pago_alumno_id   = item['pago'].id if item['pago'] else None,
        )
        db.session.add(li)

    db.session.commit()
    flash('Liquidación generada correctamente.', 'success')
    return redirect(url_for('caja.ver_liquidacion', liq_id=liq.id))
    
# ------------------------------------------------------------------
# Ver liquidación
# ------------------------------------------------------------------
@caja_bp.route('/liquidacion/<int:liq_id>')
@login_required
@rol_requerido('admin', 'profesor')
def ver_liquidacion(liq_id):
    liq = LiquidacionProfesor.query.get_or_404(liq_id)

    # Verificar que el profesor solo vea las suyas
    if current_user.es_profesor():
        if not current_user.profesor or current_user.profesor.id != liq.profesor_id:
            flash('No tenés acceso a esa liquidación.', 'danger')
            return redirect(url_for('auth.dashboard'))

    form = PagarLiquidacionForm()
    if request.method == 'GET':
        form.fecha.data = date.today()

    return render_template('caja/ver_liquidacion.html', liq=liq, form=form)


# ------------------------------------------------------------------
# Pagar liquidación
# ------------------------------------------------------------------
@caja_bp.route('/liquidacion/<int:liq_id>/pagar', methods=['POST'])
@login_required
@rol_requerido('admin')
def pagar_liquidacion(liq_id):
    liq  = LiquidacionProfesor.query.get_or_404(liq_id)
    form = PagarLiquidacionForm()

    if form.validate_on_submit():
        cuenta = _cuenta_por_tipo(form.tipo_pago.data)
        mov = _registrar_movimiento(
            cuenta_id       = cuenta.id,
            tipo            = 'egreso',
            categoria       = CAT_LIQ_PROFESOR,
            descripcion     = (f"Liquidación {liq.profesor.usuario.nombre} "
                               f"{liq.periodo_str}"),
            monto           = liq.monto_confirmado,
            fecha           = form.fecha.data,
            referencia_id   = liq.id,
            referencia_tipo = 'liquidacion',
        )
        liq.estado        = 'pagada'
        liq.fecha_pago    = form.fecha.data
        liq.tipo_pago     = form.tipo_pago.data
        liq.movimiento_id = mov.id
        db.session.commit()

        flash('Liquidación pagada correctamente.', 'success')
        return redirect(url_for('caja.comprobante_liquidacion', liq_id=liq.id))

    return redirect(url_for('caja.ver_liquidacion', liq_id=liq_id))


# ------------------------------------------------------------------
# Movimiento extraordinario
# ------------------------------------------------------------------
@caja_bp.route('/movimiento-extra', methods=['GET', 'POST'])
@login_required
@rol_requerido('admin')
def movimiento_extra():
    form = MovimientoExtraForm()
    form.categoria.choices = [
    ('cuota_alumno',               'Cuota alumno'),
    ('cobro_extra_reprogramacion', 'Cobro extra reprogramación'),
    ('ingreso_extraordinario',     'Ingreso extraordinario'),
    ('liquidacion_profesor',       'Liquidación profesor'),
    ('gasto_otro',                 'Gasto otro'),
]

    if request.method == 'GET':
        form.fecha.data = date.today()

    if form.validate_on_submit():
        cuenta = _cuenta_por_tipo(form.cuenta.data)
        _registrar_movimiento(
            cuenta_id       = cuenta.id,
            tipo            = form.tipo.data,
            categoria       = form.categoria.data,
            descripcion     = form.descripcion.data,
            monto           = form.monto.data,
            fecha           = form.fecha.data,
            referencia_tipo = 'manual',
        )
        db.session.commit()
        flash('Movimiento registrado correctamente.', 'success')
        return redirect(url_for('caja.index'))

    return render_template('caja/movimiento_extra.html', form=form)


# ------------------------------------------------------------------
# Ajuste de caja
# ------------------------------------------------------------------
@caja_bp.route('/ajuste', methods=['GET', 'POST'])
@login_required
@rol_requerido('admin')
def ajuste_caja():
    form = AjusteCajaForm()
    cuentas = CuentaCaja.query.filter_by(activo=True).all()
    form.cuenta_id.choices = [(c.id, f"{c.nombre.capitalize()} — Saldo: ${c.saldo_actual:.2f}")
                               for c in cuentas]

    if request.method == 'GET':
        form.fecha.data = date.today()

    if form.validate_on_submit():
        cuenta = CuentaCaja.query.get(form.cuenta_id.data)
        ajuste = AjusteCaja(
            cuenta_id      = form.cuenta_id.data,
            monto_anterior = cuenta.saldo_actual,
            monto_nuevo    = form.monto_nuevo.data,
            motivo         = form.motivo.data,
            fecha          = form.fecha.data,
            creado_por_id  = current_user.id,
        )
        # Ajustar saldo_inicial para que saldo_actual == monto_nuevo
        diferencia = float(form.monto_nuevo.data) - cuenta.saldo_actual
        cuenta.saldo_inicial = float(cuenta.saldo_inicial) + diferencia
        db.session.add(ajuste)
        db.session.commit()
        flash(f'Saldo de {cuenta.nombre} ajustado correctamente.', 'success')
        return redirect(url_for('caja.index'))

    return render_template('caja/ajuste_caja.html', form=form, cuentas=cuentas)


# ------------------------------------------------------------------
# Buscador de movimientos
# ------------------------------------------------------------------
@caja_bp.route('/movimientos')
@login_required
@rol_requerido('admin')
def movimientos():
    hoy    = date.today()
    q      = request.args.get('q', '').strip()
    mes    = request.args.get('mes',  type=int)
    anio   = request.args.get('anio', type=int)
    tipo   = request.args.get('tipo', '')
    cuenta = request.args.get('cuenta', '')

    query = MovimientoCaja.query

    if q:
        query = query.filter(MovimientoCaja.descripcion.ilike(f'%{q}%'))
    if mes and anio:
        query = query.filter(extract('month', MovimientoCaja.fecha) == mes)
        query = query.filter(extract('year',  MovimientoCaja.fecha) == anio)
    elif anio:
        query = query.filter(extract('year', MovimientoCaja.fecha) == anio)
    if tipo:
        query = query.filter_by(tipo=tipo)
    if cuenta:
        cuenta_obj = CuentaCaja.query.filter_by(nombre=cuenta).first()
        if cuenta_obj:
            query = query.filter_by(cuenta_id=cuenta_obj.id)

    movimientos = query.order_by(MovimientoCaja.fecha.desc(),
                                 MovimientoCaja.creado_en.desc()).all()

    total_ingresos = sum(float(m.monto) for m in movimientos if m.tipo == 'ingreso')
    total_egresos  = sum(float(m.monto) for m in movimientos if m.tipo == 'egreso')

    return render_template('caja/movimientos.html',
                           movimientos=movimientos,
                           total_ingresos=total_ingresos,
                           total_egresos=total_egresos,
                           q=q, mes=mes, anio=anio, tipo=tipo, cuenta=cuenta,
                           hoy=hoy)


# ------------------------------------------------------------------
# Exportar movimientos a Excel
# ------------------------------------------------------------------
@caja_bp.route('/movimientos/exportar')
@login_required
@rol_requerido('admin')
def exportar_movimientos():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        q      = request.args.get('q', '').strip()
        mes    = request.args.get('mes',  type=int)
        anio   = request.args.get('anio', type=int)
        tipo   = request.args.get('tipo', '')
        cuenta = request.args.get('cuenta', '')

        query = MovimientoCaja.query
        if q:
            query = query.filter(MovimientoCaja.descripcion.ilike(f'%{q}%'))
        if mes and anio:
            query = query.filter(extract('month', MovimientoCaja.fecha) == mes)
            query = query.filter(extract('year',  MovimientoCaja.fecha) == anio)
        if tipo:
            query = query.filter_by(tipo=tipo)
        if cuenta:
            cuenta_obj = CuentaCaja.query.filter_by(nombre=cuenta).first()
            if cuenta_obj:
                query = query.filter_by(cuenta_id=cuenta_obj.id)

        movs = query.order_by(MovimientoCaja.fecha.desc()).all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Movimientos'

        headers = ['Fecha','Tipo','Categoría','Descripción','Cuenta','Monto']
        header_fill = PatternFill('solid', fgColor='5C1A1A')
        header_font = Font(color='F5E6DA', bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row, m in enumerate(movs, 2):
            ws.cell(row=row, column=1, value=m.fecha.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=m.tipo.capitalize())
            ws.cell(row=row, column=3, value=m.categoria_label)
            ws.cell(row=row, column=4, value=m.descripcion)
            ws.cell(row=row, column=5, value=m.cuenta.nombre.capitalize())
            ws.cell(row=row, column=6, value=float(m.monto))

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"movimientos_eiren_{date.today().isoformat()}.xlsx"
        return send_file(output, as_attachment=True,
                         download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except ImportError:
        flash('Necesitás instalar openpyxl: pip install openpyxl', 'danger')
        return redirect(url_for('caja.movimientos'))


# ------------------------------------------------------------------
# Comprobante pago alumno (HTML imprimible)
# ------------------------------------------------------------------
@caja_bp.route('/comprobante/pago/<int:pago_id>')
@login_required
@rol_requerido('admin')
def comprobante_pago(pago_id):
    pago = PagoAlumno.query.get_or_404(pago_id)
    return render_template('caja/comprobante_pago.html', pago=pago)


# ------------------------------------------------------------------
# Comprobante liquidación (HTML imprimible)
# ------------------------------------------------------------------
@caja_bp.route('/comprobante/liquidacion/<int:liq_id>')
@login_required
@rol_requerido('admin', 'profesor')
def comprobante_liquidacion(liq_id):
    liq = LiquidacionProfesor.query.get_or_404(liq_id)
    if current_user.es_profesor():
        if not current_user.profesor or current_user.profesor.id != liq.profesor_id:
            flash('No tenés acceso a esa liquidación.', 'danger')
            return redirect(url_for('auth.dashboard'))
    return render_template('caja/comprobante_liquidacion.html', liq=liq)
